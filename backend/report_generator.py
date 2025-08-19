"""报告生成模块"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from database import Task, Issue
import os

def generate_report(task_id: int, db: Session) -> str:
    """生成Excel报告 - 包含所有问题详细信息"""
    # 获取任务和问题
    task = db.query(Task).filter(Task.id == task_id).first()
    if not task:
        raise ValueError(f"任务不存在: {task_id}")
    
    issues = db.query(Issue).filter(Issue.task_id == task_id).all()
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "问题汇总"
    
    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # 写入表头 - 包含所有字段
    headers = [
        '序号', '问题类型', '详细描述', '所在位置', '严重程度', 
        '置信度', '原文内容', '修改建议', '用户影响', '判定理由', 
        '上下文', '用户反馈', '用户评价'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    
    # 设置行高
    ws.row_dimensions[1].height = 30
    
    # 写入数据
    for i, issue in enumerate(issues, 2):
        ws.cell(row=i, column=1, value=i-1)
        ws.cell(row=i, column=2, value=issue.issue_type)
        ws.cell(row=i, column=3, value=issue.description).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=4, value=issue.location)
        ws.cell(row=i, column=5, value=issue.severity)
        ws.cell(row=i, column=6, value=f"{(issue.confidence * 100):.0f}%" if issue.confidence else "")
        ws.cell(row=i, column=7, value=issue.original_text or "").alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=8, value=issue.suggestion).alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=9, value=issue.user_impact or "").alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=10, value=issue.reasoning or "").alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=11, value=issue.context or "").alignment = Alignment(wrap_text=True)
        ws.cell(row=i, column=12, value=issue.feedback_type or '未反馈')
        ws.cell(row=i, column=13, value=issue.feedback_comment or '')
    
    # 调整列宽 - 根据内容自适应
    column_widths = [8, 15, 50, 25, 12, 10, 40, 50, 40, 40, 30, 12, 30]
    for i, width in enumerate(column_widths, 1):
        if i <= len(column_widths):
            ws.column_dimensions[chr(64 + i)].width = width
    
    # 添加统计信息
    stats_row = len(issues) + 3
    ws.cell(row=stats_row, column=1, value='统计信息').font = Font(bold=True)
    ws.cell(row=stats_row + 1, column=1, value='总问题数')
    ws.cell(row=stats_row + 1, column=2, value=len(issues))
    
    # 按严重程度统计
    severity_counts = {}
    for issue in issues:
        severity_counts[issue.severity] = severity_counts.get(issue.severity, 0) + 1
    
    row_offset = 2
    for severity in ['致命', '严重', '一般', '提示']:
        if severity in severity_counts:
            ws.cell(row=stats_row + row_offset, column=1, value=f'{severity}级别')
            ws.cell(row=stats_row + row_offset, column=2, value=severity_counts[severity])
            row_offset += 1
    
    # 用户反馈统计
    accepted = len([i for i in issues if i.feedback_type == 'accept'])
    rejected = len([i for i in issues if i.feedback_type == 'reject'])
    
    ws.cell(row=stats_row + row_offset, column=1, value='已接受')
    ws.cell(row=stats_row + row_offset, column=2, value=accepted)
    ws.cell(row=stats_row + row_offset + 1, column=1, value='已拒绝')
    ws.cell(row=stats_row + row_offset + 1, column=2, value=rejected)
    ws.cell(row=stats_row + row_offset + 2, column=1, value='未反馈')
    ws.cell(row=stats_row + row_offset + 2, column=2, value=len(issues) - accepted - rejected)
    
    # 添加任务信息
    task_info_row = stats_row + row_offset + 4
    ws.cell(row=task_info_row, column=1, value='任务信息').font = Font(bold=True)
    ws.cell(row=task_info_row + 1, column=1, value='文件名')
    ws.cell(row=task_info_row + 1, column=2, value=task.file_name)
    ws.cell(row=task_info_row + 2, column=1, value='检测时间')
    ws.cell(row=task_info_row + 2, column=2, value=task.created_at.strftime('%Y-%m-%d %H:%M:%S') if task.created_at else '')
    ws.cell(row=task_info_row + 3, column=1, value='完成时间')
    ws.cell(row=task_info_row + 3, column=2, value=task.completed_at.strftime('%Y-%m-%d %H:%M:%S') if task.completed_at else '')
    
    # 保存文件
    os.makedirs("./data/reports", exist_ok=True)
    file_name = f"{task_id}_{task.file_name.replace('.', '_')}_report.xlsx"
    file_path = f"./data/reports/{file_name}"
    wb.save(file_path)
    
    return file_path